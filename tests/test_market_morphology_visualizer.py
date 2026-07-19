from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from market_engine.market_morphology_visualizer_cli import export_visual_comparison


def _frame(multiplier: float = 1.0) -> pd.DataFrame:
    dates = pd.date_range("2026-01-01", periods=8, freq="D")
    totals = [100, 105, 110, 108, 102, 106, 111, 115]
    rows: list[dict[str, object]] = []
    for date, total in zip(dates, totals):
        rows.extend(
            [
                {"date": date, "ticker": "AAA", "market_cap": total * 0.6 * multiplier},
                {"date": date, "ticker": "BBB", "market_cap": total * 0.4 * multiplier},
            ]
        )
    return pd.DataFrame(rows)


def test_visual_export_creates_expected_sheets_and_charts(tmp_path: Path) -> None:
    output = tmp_path / "visual.xlsx"

    export_visual_comparison(
        _frame(),
        _frame(multiplier=3.0),
        output,
        smoothing_window=3,
        rolling_window=3,
    )

    assert output.exists()
    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Resumen",
        "Metricas",
        "Comparacion_Visual",
        "Fidelidad_Movil",
    ]
    assert len(workbook["Comparacion_Visual"]._charts) == 4


def test_visual_export_uses_requested_labels(tmp_path: Path) -> None:
    output = tmp_path / "labels.xlsx"

    export_visual_comparison(
        _frame(),
        _frame(multiplier=2.0),
        output,
        smoothing_window=2,
        rolling_window=3,
        sample_label="MUESTRA",
        reference_label="REFERENCIA",
    )

    workbook = load_workbook(output, read_only=False)
    headers = [cell.value for cell in workbook["Comparacion_Visual"][1]]
    assert "MUESTRA_index" in headers
    assert "REFERENCIA_index" in headers
    assert "MUESTRA_drawdown_pct" in headers
    assert "REFERENCIA_drawdown_pct" in headers
